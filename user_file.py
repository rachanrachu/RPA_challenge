import openpyxl
from openpyxl import *

wb=load_workbook("TA - RPA Challenge Shopping List.xlsx")

def getrowcount(filename,sheetname):
    excel_data = openpyxl.load_workbook(filename)
    excel_sheet = excel_data.get_sheet_by_name(sheetname)
    return (excel_sheet.max_row)


def getcoloumcount(filename,sheetname):
    excel_data = openpyxl.load_workbook(filename)
    excel_sheet = excel_data.get_sheet_by_name(sheetname)
    return (excel_sheet.max_column)

def readdatafile(filename,sheetname,rownum,columnno):
    excel_data = openpyxl.load_workbook(filename)
    excel_sheet = excel_data.get_sheet_by_name(sheetname)
    return excel_sheet.cell(row=rownum, column=columnno).value

def writedatafile(write_data,filename,sheetname,rownum,columnno):
    ws = wb[sheetname]
    wcell1 = ws.cell(rownum, columnno)
    wcell1.value = write_data
    wb.save(filename)

